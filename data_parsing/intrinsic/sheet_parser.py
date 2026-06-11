"""Parse parameter blocks from intrinsic-properties Excel sheets."""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from .config import DRUG_SHEETS, METADATA_ROW_LABELS, MORPHOLOGY_ROW_LABELS, SKIP_PARAM_LABELS
from .ids import HeaderInfo, normalize_id, parse_headers_row
from .labels import canonical_label, iv_header_label, normalize_section


@dataclass
class ParsedValue:
    cell_id: str
    source_sheet: str
    block: str  # control | effect
    section: str
    label: str
    value: Any
    col: int
    row: int


@dataclass
class SheetMeta:
    classic_burster: dict[str, Any] = field(default_factory=dict)
    area_morph_raw: dict[str, Any] = field(default_factory=dict)


@dataclass
class SheetParseResult:
    sheet_name: str
    headers: list[HeaderInfo]
    values: list[ParsedValue]
    meta: SheetMeta
    task_exclude: set[str] = field(default_factory=set)
    task_notes: dict[str, str] = field(default_factory=dict)
    cesium_secondary_skipped: dict[str, str] = field(default_factory=dict)


def _is_id_header(val: Any) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return s.lower().startswith("nm") or bool(__import__("re").match(r"^\d{4}_\d{2}_\d{2}_c\d+", s))


def _detect_headers(ws: Worksheet, sheet_name: str) -> list[HeaderInfo]:
    if sheet_name == "V2M_L5_TASK_Acidic_pH":
        pairs = [(c, ws.cell(1, c).value) for c in range(2, ws.max_column + 1)]
        return parse_headers_row(pairs, disambiguate_dupes=True)
    if sheet_name == "V2M_L5_Caesum":
        pairs = [(c, ws.cell(1, c).value) for c in range(4, ws.max_column + 1)]
        return parse_headers_row(pairs)
    pairs = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if _is_id_header(v):
            pairs.append((c, v))
    return parse_headers_row(pairs)


def _read_meta_rows(ws: Worksheet, headers: list[HeaderInfo]) -> SheetMeta:
    meta = SheetMeta()
    id_by_col = {h.col: h.normalized for h in headers}
    for r in range(1, min(ws.max_row, 20) + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        ls = str(label).strip()
        if ls in METADATA_ROW_LABELS:
            key = METADATA_ROW_LABELS[ls]
            for col, cid in id_by_col.items():
                meta.classic_burster.setdefault(cid, None)
                if key == "classic_burster":
                    meta.classic_burster[cid] = ws.cell(r, col).value
        if ls in MORPHOLOGY_ROW_LABELS:
            for col, cid in id_by_col.items():
                meta.area_morph_raw[cid] = ws.cell(r, col).value
    return meta


def _mega_block_from_header(header: str) -> str | None:
    if "_IV" not in header:
        return None
    if "CONTROL" in header:
        return "control"
    if header.strip().startswith("_IV"):
        return "effect"
    return None


def _parse_block_sheet(
    ws: Worksheet,
    sheet_name: str,
    *,
    blocks: tuple[str, ...],
    merge_map: dict[tuple[str, str], str],
) -> SheetParseResult:
    headers = _detect_headers(ws, sheet_name)
    meta = _read_meta_rows(ws, headers)
    values: list[ParsedValue] = []
    id_by_col = {h.col: h.normalized for h in headers}

    mega_block: str | None = None
    section: str | None = None
    pending_condition: dict[int, Any] = {}

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        sa = str(a).strip()

        if sa.startswith("_"):
            mb = _mega_block_from_header(sa)
            if mb:
                mega_block = mb
            section = normalize_section(sa)
            pending_condition = {}
            iv_label = iv_header_label(sa)
            if iv_label and section == "_IV" and mega_block in blocks:
                canon = canonical_label(section, iv_label, merge_map)
                if canon:
                    for col, cid in id_by_col.items():
                        val = ws.cell(r, col).value
                        values.append(
                            ParsedValue(
                                cell_id=cid,
                                source_sheet=sheet_name,
                                block=mega_block,
                                section=section,
                                label=canon,
                                value=val,
                                col=col,
                                row=r,
                            )
                        )
            continue

        if sa == "condition":
            pending_condition = {col: ws.cell(r, col).value for col in id_by_col}
            continue

        if section is None or mega_block not in blocks:
            continue
        if sa in SKIP_PARAM_LABELS:
            continue

        canon = canonical_label(section, sa, merge_map)
        if not canon:
            continue

        for col, cid in id_by_col.items():
            cond = pending_condition.get(col)
            # Spec: 0=control, 1=effect, 2/Washout/empty ignored
            if cond not in (0, 0.0, 1, 1.0):
                continue
            if mega_block == "control" and cond not in (0, 0.0):
                continue
            if mega_block == "effect" and cond not in (1, 1.0):
                continue
            val = ws.cell(r, col).value
            values.append(
                ParsedValue(
                    cell_id=cid,
                    source_sheet=sheet_name,
                    block=mega_block,
                    section=section,
                    label=canon,
                    value=val,
                    col=col,
                    row=r,
                )
            )

    return SheetParseResult(sheet_name=sheet_name, headers=headers, values=values, meta=meta)


def parse_standard_sheet(ws: Worksheet, sheet_name: str, merge_map: dict) -> SheetParseResult:
    return _parse_block_sheet(ws, sheet_name, blocks=("control",), merge_map=merge_map)


def parse_drug_sheet(ws: Worksheet, sheet_name: str, merge_map: dict) -> SheetParseResult:
    return _parse_block_sheet(ws, sheet_name, blocks=("control", "effect"), merge_map=merge_map)


def parse_task_sheet(ws: Worksheet, merge_map: dict) -> SheetParseResult:
    sheet_name = "V2M_L5_TASK_Acidic_pH"
    headers = _detect_headers(ws, sheet_name)
    meta = SheetMeta()
    values: list[ParsedValue] = []
    id_by_col = {h.col: h.normalized for h in headers}

    task_exclude: set[str] = set()
    for r in range(1, 10):
        a = ws.cell(r, 1).value
        if a and "exlu" in str(a).lower():
            for col, cid in id_by_col.items():
                flag = ws.cell(r, col).value
                if flag in (1, 1.0, "1"):
                    task_exclude.add(cid)

    current_block: str | None = None
    section: str | None = None
    control_chirp_labels: list[str] = []

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        sa = str(a).strip()

        if sa == "pH 7.3":
            current_block = "control"
            section = None
            continue
        if sa == "pH 6.1":
            current_block = "effect"
            section = None
            continue

        if sa.startswith("_"):
            section = normalize_section(sa)
            if section == "_chirp" and current_block == "control":
                control_chirp_labels = []
            continue

        if sa == "condition":
            continue

        if current_block is None:
            continue

        # pH 6.1 chirp block lacks _chirp header — infer from control chirp labels
        if current_block == "effect" and section is None:
            if sa in control_chirp_labels or sa in (
                "Vm (avg mV)",
                "Ihold (avg pA)",
                "Res. freq. (Hz)",
                "Res. imp. mag. (MOhm)",
                "Trough impedance (MOhm)",
                "Relative imp. mag.",
                "Synchronous Frequency (Hz)",
                "Phase lead integral (rad*Hz)",
            ):
                section = "_chirp"

        if section is None:
            continue

        canon = canonical_label(section, sa, merge_map)
        if not canon:
            continue

        if section == "_chirp" and current_block == "control":
            control_chirp_labels.append(canon)

        for col, cid in id_by_col.items():
            if cid in task_exclude:
                continue
            val = ws.cell(r, col).value
            values.append(
                ParsedValue(
                    cell_id=cid,
                    source_sheet=sheet_name,
                    block=current_block,
                    section=section,
                    label=canon,
                    value=val,
                    col=col,
                    row=r,
                )
            )

    task_notes = {h.normalized: h.task_note for h in headers if h.task_note}
    return SheetParseResult(
        sheet_name=sheet_name,
        headers=headers,
        values=values,
        meta=meta,
        task_exclude=task_exclude,
        task_notes=task_notes,
    )


def parse_cesium_sheet(ws: Worksheet, merge_map: dict) -> SheetParseResult:
    sheet_name = "V2M_L5_Caesum"
    headers = _detect_headers(ws, sheet_name)
    meta = SheetMeta()
    values: list[ParsedValue] = []
    id_by_col = {h.col: h.normalized for h in headers}
    cesium_skipped: dict[str, str] = {}

    section: str | None = None
    in_secondary = False

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        sa = str(a).strip()

        if sa.startswith("_"):
            section = normalize_section(sa)
            in_secondary = False
            continue

        if sa == "condition":
            # Secondary block: condition row with only cols G-H populated at ~row 17
            row_vals = [ws.cell(r, c).value for c in id_by_col]
            if any(v is not None for v in row_vals):
                # Detect sparse secondary (−56 mV) block
                filled = sum(1 for c in id_by_col if ws.cell(r, c).value is not None)
                if filled <= 2 and r > 15:
                    in_secondary = True
                    for col, cid in id_by_col.items():
                        if ws.cell(r, col).value is not None:
                            cesium_skipped[cid] = "ambiguous secondary −56 mV block; primary block used"
            continue

        if section is None or in_secondary:
            continue

        canon = canonical_label(section, sa, merge_map)
        if not canon:
            continue

        for col, cid in id_by_col.items():
            val = ws.cell(r, col).value
            values.append(
                ParsedValue(
                    cell_id=cid,
                    source_sheet=sheet_name,
                    block="effect",
                    section=section,
                    label=canon,
                    value=val,
                    col=col,
                    row=r,
                )
            )

    return SheetParseResult(
        sheet_name=sheet_name,
        headers=headers,
        values=values,
        meta=meta,
        cesium_secondary_skipped=cesium_skipped,
    )


parse_caesum_sheet = parse_cesium_sheet  # Excel sheet name retains student typo


def sheet_region_layer(sheet_name: str) -> tuple[str, str]:
    region = ""
    layer = ""
    if sheet_name.startswith("V1"):
        region = "V1"
    elif sheet_name.startswith("V2M") or sheet_name.startswith("All"):
        if "V2M" in sheet_name or sheet_name == "All Analysed data":
            region = "V2M" if sheet_name != "All Analysed data" else ""
    if "L2-3" in sheet_name or "L2/3" in sheet_name:
        layer = "L2-3"
    elif "L5" in sheet_name:
        layer = "L5"
    return region, layer
