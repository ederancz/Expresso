"""Phase 2 integrity checks — run before writing outputs."""

from __future__ import annotations

import re
from typing import Any

from openpyxl import Workbook

from .config import DRUG_SHEETS, PARAMETER_DATA_SHEETS
from .ids import normalize_id
from .labels import col_name
from .sheet_parser import ParsedValue, SheetParseResult

CESIUM_EXPERIMENT = DRUG_SHEETS["V2M_L5_Caesum"]
MAX_BYTE_EQUAL_ERRORS = 20


class IntegrityCheckError(Exception):
    """Raised when Phase 2 pre-write checks fail."""

    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__(self.summary())

    def summary(self) -> str:
        head = f"Phase 2 integrity check failed ({len(self.errors)} error(s))"
        if not self.errors:
            return head
        shown = self.errors[:MAX_BYTE_EQUAL_ERRORS]
        tail = ""
        if len(self.errors) > MAX_BYTE_EQUAL_ERRORS:
            tail = f"\n  … and {len(self.errors) - MAX_BYTE_EQUAL_ERRORS} more"
        return head + ":\n  " + "\n  ".join(shown) + tail


def byte_equal(source: Any, copied: Any) -> bool:
    """Strict equality — no coercion or rounding."""
    return source == copied


def verify_parsed_byte_equality(
    wb: Workbook,
    parse_results: list[SheetParseResult],
) -> list[str]:
    """Re-read every parsed cell from the workbook and assert it matches."""
    errors: list[str] = []
    for pr in parse_results:
        ws = wb[pr.sheet_name]
        for pv in pr.values:
            reread = ws.cell(pv.row, pv.col).value
            if not byte_equal(reread, pv.value):
                errors.append(
                    f"byte-equality: {pr.sheet_name} r{pv.row}c{pv.col} "
                    f"cell={pv.cell_id!r} block={pv.block} "
                    f"{col_name(pv.section, pv.label)!r}: "
                    f"source={reread!r} copied={pv.value!r}"
                )
    return errors


def verify_pharmacology_subset(
    control_ids: set[str],
    effect_rows: list[dict[str, Any]],
) -> tuple[list[str], list[str]]:
    """
    pharmacology_effect IDs must be subset of control_excitability IDs,
    except cesium (effect-only) cells.
    Returns (errors, cesium_only_cell_ids).
    """
    errors: list[str] = []
    cesium_only: set[str] = set()
    for row in effect_rows:
        cid = row["cell_id"]
        if cid in control_ids:
            continue
        if row.get("experiment") == CESIUM_EXPERIMENT:
            cesium_only.add(cid)
        else:
            errors.append(
                f"pharmacology subset: {cid!r} in pharmacology_effect "
                f"({row.get('experiment')!r}) but not in control_excitability "
                f"(non-cesium violation)"
            )
    return errors, sorted(cesium_only)


def count_orphan_ids(wb: Workbook, included_ids: set[str]) -> tuple[int, list[str]]:
    """
    IDs on ALL_CELLS_NEW absent from all parameter sheets (informational).
    Orphan rescue is not implemented; count should be 0 for the known dataset.
    """
    if "ALL_CELLS_NEW" not in wb.sheetnames:
        return 0, []
    ws = wb["ALL_CELLS_NEW"]
    orphans: set[str] = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s.lower().startswith("nm") or re.match(r"^\d{4}_", s):
                nid = normalize_id(s)
                if nid not in included_ids:
                    orphans.add(nid)
    return len(orphans), sorted(orphans)


def _all_included_ids(parse_results: list[SheetParseResult]) -> set[str]:
    ids: set[str] = set()
    for pr in parse_results:
        for h in pr.headers:
            ids.add(h.normalized)
    return ids


def run_phase2_checks(
    wb: Workbook,
    *,
    parse_results: list[SheetParseResult],
    control_ids: set[str],
    effect_rows: list[dict[str, Any]],
    label_merge_count: int,
    conflict_cell_count: int,
    cluster_fill_count: int,
    excluded_in_may_count: int,
    exclude_flag_dropped: int = 0,
) -> dict[str, Any]:
    """
    Run all Phase 2 checks. Raises IntegrityCheckError on failure.
    Returns a report dict recorded in run_manifest.json.
    """
    errors: list[str] = []
    errors.extend(verify_parsed_byte_equality(wb, parse_results))

    subset_errors, cesium_only = verify_pharmacology_subset(control_ids, effect_rows)
    errors.extend(subset_errors)

    orphan_count, orphan_ids = count_orphan_ids(wb, _all_included_ids(parse_results))

    values_checked = sum(len(pr.values) for pr in parse_results)

    report: dict[str, Any] = {
        "passed": len(errors) == 0,
        "values_byte_checked": values_checked,
        "control_neurons": len(control_ids),
        "pharmacology_effect_rows": len(effect_rows),
        "duplicate_conflict_cells": conflict_cell_count,
        "label_merges": label_merge_count,
        "orphan_ids_found": orphan_count,
        "orphan_ids": orphan_ids,
        "cluster_fills_assumed_type": cluster_fill_count,
        "cluster_fill_column": "assumed_type",
        "excluded_in_may_dropped": excluded_in_may_count,
        "exclude_flag_dropped": exclude_flag_dropped,
        "cesium_effect_only_cells": cesium_only,
        "errors": errors,
    }

    if errors:
        raise IntegrityCheckError(errors)

    return report
