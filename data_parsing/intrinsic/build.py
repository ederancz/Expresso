"""Phase 1 orchestrator — build Intrinsic_master outputs."""

from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook

from .config import (
    DEFAULT_OUTPUT_DIR,
    DEFAULT_SOURCE,
    DRUG_SHEETS,
    PARAMETER_DATA_SHEETS,
    STANDARD_SHEETS,
)
from .dedup import deduplicate_control_instances, merge_sheet_metas, pick_canonical
from .ids import duplicate_header_warnings
from .labels import build_label_merge_map, col_name
from .manifest import write_run_manifest
from .metadata import (
    fill_assumed_type_from_cluster,
    load_all_cells,
    load_cluster_metadata,
    load_exclude_flag_ids,
    load_excluded_may_ids,
    load_metadata_tags,
    merge_cell_metadata,
    metadata_to_row,
)
from .sheet_parser import (
    ParsedValue,
    SheetParseResult,
    parse_cesium_sheet,
    parse_drug_sheet,
    parse_standard_sheet,
    parse_task_sheet,
    sheet_region_layer,
)
from .values import format_param_for_output
from .qc_outliers import write_qc_workbook
from .verify import IntegrityCheckError, run_phase2_checks
from .verify_postbuild import PostBuildVerificationError, run_phase3_checks

META_COLUMNS = [
    "cell_id",
    "region",
    "areaCCF",
    "layer",
    "projection_target",
    "assumed_type",
    "physiological_cluster",
    "source_sheet",
    "classic_burster",
    "exclude_flag",
    "cre_label",
    "axon",
    "notes",
    "time_from_5HT",
    "dup_conflict",
]

CONFLICT_META_COLUMNS = [
    "cell_id",
    "region",
    "areaCCF",
    "layer",
    "projection_target",
    "assumed_type",
    "physiological_cluster",
    "source_sheet",
    "conflict_source_sheet",
]

OUTPUT_SIGFIGS = 4


def _collect_raw_label_pairs(parse_results: list[SheetParseResult]) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for pr in parse_results:
        for pv in pr.values:
            pairs.add((pv.section, pv.label))
    return pairs


def _params_from_values(values: list[ParsedValue]) -> dict[str, Any]:
    params: dict[str, Any] = {}
    for pv in values:
        key = col_name(pv.section, pv.label)
        params[key] = pv.value
    return params


def _sheet_meta_bundle(pr: SheetParseResult) -> dict[str, dict]:
    return {
        "classic_burster": pr.meta.classic_burster,
        "area_morph_raw": pr.meta.area_morph_raw,
    }


def _parse_all_sheets(wb: Workbook, merge_map: dict) -> list[SheetParseResult]:
    results: list[SheetParseResult] = []
    for name in PARAMETER_DATA_SHEETS:
        ws = wb[name]
        if name in STANDARD_SHEETS:
            results.append(parse_standard_sheet(ws, name, merge_map))
        elif name == "V2M_L5_TASK_Acidic_pH":
            results.append(parse_task_sheet(ws, merge_map))
        elif name == "V2M_L5_Caesum":
            results.append(parse_cesium_sheet(ws, merge_map))
        else:
            results.append(parse_drug_sheet(ws, name, merge_map))
    return results


def _collect_header_warnings(parse_results: list[SheetParseResult]) -> list[str]:
    out: list[str] = []
    for pr in parse_results:
        out.extend(duplicate_header_warnings(pr.headers))
    return out


def _group_control(
    parse_results: list[SheetParseResult],
    dropped_ids: set[str],
) -> dict[str, list[dict[str, Any]]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for pr in parse_results:
        sheet_name = pr.sheet_name
        if sheet_name == "V2M_L5_Caesum":
            continue

        by_cell: dict[str, list[ParsedValue]] = defaultdict(list)
        for pv in pr.values:
            if pv.block != "control":
                continue
            by_cell[pv.cell_id].append(pv)

        region, layer = sheet_region_layer(sheet_name)
        meta_bundle = _sheet_meta_bundle(pr)
        for cid, pvs in by_cell.items():
            if cid in pr.task_exclude or cid in dropped_ids:
                continue
            groups[cid].append(
                {
                    "cell_id": cid,
                    "source_sheet": sheet_name,
                    "region": region,
                    "layer": layer,
                    "params": _params_from_values(pvs),
                    "sheet_meta": meta_bundle,
                }
            )
    return groups


def _group_effect(
    parse_results: list[SheetParseResult],
    dropped_ids: set[str],
) -> dict[tuple[str, str], list[dict[str, Any]]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for pr in parse_results:
        sheet_name = pr.sheet_name
        if sheet_name not in DRUG_SHEETS:
            continue
        experiment = DRUG_SHEETS[sheet_name]

        by_cell: dict[str, list[ParsedValue]] = defaultdict(list)
        for pv in pr.values:
            if pv.block != "effect":
                continue
            by_cell[pv.cell_id].append(pv)

        region, layer = sheet_region_layer(sheet_name)
        meta_bundle = _sheet_meta_bundle(pr)
        for cid, pvs in by_cell.items():
            if cid in pr.task_exclude or cid in dropped_ids:
                continue
            groups[(cid, experiment)].append(
                {
                    "cell_id": cid,
                    "source_sheet": sheet_name,
                    "experiment": experiment,
                    "region": region,
                    "layer": layer,
                    "params": _params_from_values(pvs),
                    "sheet_meta": meta_bundle,
                }
            )
    return groups


def _ordered_param_columns(all_params: set[str]) -> list[str]:
    section_order = ["_IV", "_short_depol", "_EPSP", "_crit_freq", "_sag", "_chirp"]
    by_sec: dict[str, list[str]] = defaultdict(list)
    for p in all_params:
        sec = p.split("__", 1)[0]
        by_sec[sec].append(p)
    out: list[str] = []
    for sec in section_order:
        out.extend(sorted(by_sec.get(sec, [])))
    for sec in sorted(by_sec):
        if sec not in section_order:
            out.extend(sorted(by_sec[sec]))
    return out


def _conflict_param_keys(pair_details: list[dict[str, Any]]) -> set[str]:
    keys: set[str] = set()
    for pair in pair_details:
        keys.update(pair.get("conflicting_params", {}).keys())
    return keys


def _sparse_conflict_params(
    params: dict[str, Any],
    conflict_keys: set[str],
) -> dict[str, Any]:
    return {k: params[k] for k in conflict_keys if k in params}


def _format_all_row_params(
    rows: list[dict[str, Any]],
    *,
    skip_keys: set[str],
) -> None:
    for row in rows:
        for k in list(row.keys()):
            if k not in skip_keys:
                row[k] = format_param_for_output(row[k], OUTPUT_SIGFIGS)


def build_master(
    source_path: Path = DEFAULT_SOURCE,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
) -> dict[str, Any]:
    wb = openpyxl.load_workbook(source_path, data_only=True, read_only=False)

    preliminary = _parse_all_sheets(wb, merge_map={})
    raw_pairs = _collect_raw_label_pairs(preliminary)
    merge_map, merge_log = build_label_merge_map(raw_pairs)

    parse_results = _parse_all_sheets(wb, merge_map)
    header_warnings = _collect_header_warnings(parse_results)

    all_cells = load_all_cells(wb)
    excluded_may = load_excluded_may_ids(wb)
    exclude_flag = load_exclude_flag_ids(wb)
    dropped_ids = excluded_may | exclude_flag
    tags = load_metadata_tags(wb)
    clusters = load_cluster_metadata(wb)

    control_groups = _group_control(parse_results, dropped_ids)
    effect_groups = _group_effect(parse_results, dropped_ids)

    control_rows: list[dict[str, Any]] = []
    conflict_rows: list[dict[str, Any]] = []
    conflicts_detail: list[dict[str, Any]] = []
    region_conflicts: list[dict[str, Any]] = []
    cluster_fills = 0
    all_conflict_param_cols: set[str] = set()

    for cid, instances in sorted(control_groups.items()):
        merged_meta = merge_sheet_metas(instances)
        canonical, conflicts, has_conflict, pair_details = deduplicate_control_instances(instances)
        if canonical is None:
            continue
        if len(instances) > 1:
            canonical = pick_canonical(instances)
        canonical["sheet_meta"] = merged_meta

        meta = merge_cell_metadata(
            cid,
            source_sheet=canonical["source_sheet"],
            region=canonical["region"],
            layer=canonical["layer"],
            sheet_meta=canonical["sheet_meta"],
            all_cells=all_cells,
            tags=tags,
            clusters=clusters,
        )
        if fill_assumed_type_from_cluster(meta):
            cluster_fills += 1

        if meta.region_conflict:
            region_conflicts.append(
                {
                    "cell_id": cid,
                    "region_from_sheet": meta.region_sheet,
                    "region_from_metadata_ccf": meta.region,
                    "areaCCF": meta.area_ccf,
                    "source_sheet": meta.source_sheet,
                }
            )

        meta.dup_conflict = has_conflict
        row = metadata_to_row(meta)
        row["cell_id"] = cid
        row.update(canonical["params"])
        control_rows.append(row)

        if has_conflict:
            cell_conflict_keys = _conflict_param_keys(pair_details)
            all_conflict_param_cols |= cell_conflict_keys
            conflicts_detail.append(
                {
                    "cell_id": cid,
                    "conflicting_params": sorted(cell_conflict_keys),
                    "pair_conflicts": pair_details,
                }
            )
            for inst in conflicts:
                cm = merge_cell_metadata(
                    cid,
                    source_sheet=inst["source_sheet"],
                    region=inst["region"],
                    layer=inst["layer"],
                    sheet_meta=inst["sheet_meta"],
                    all_cells=all_cells,
                    tags=tags,
                    clusters=clusters,
                )
                crow = metadata_to_row(cm)
                crow["cell_id"] = cid
                crow["conflict_source_sheet"] = inst["source_sheet"]
                crow.update(_sparse_conflict_params(inst["params"], cell_conflict_keys))
                conflict_rows.append(crow)

    effect_rows: list[dict[str, Any]] = []
    for (cid, experiment), instances in sorted(effect_groups.items()):
        inst = instances[0] if len(instances) == 1 else pick_canonical(instances)
        meta = merge_cell_metadata(
            cid,
            source_sheet=inst["source_sheet"],
            region=inst["region"],
            layer=inst["layer"],
            sheet_meta=inst["sheet_meta"],
            all_cells=all_cells,
            tags=tags,
            clusters=clusters,
        )
        fill_assumed_type_from_cluster(meta)
        row = metadata_to_row(meta)
        row["cell_id"] = cid
        row["experiment"] = experiment
        row.update(inst["params"])
        effect_rows.append(row)

    all_param_cols: set[str] = set()
    meta_set = set(META_COLUMNS)
    for r in control_rows + effect_rows:
        all_param_cols.update(k for k in r if k not in meta_set and k != "experiment")
    param_cols = _ordered_param_columns(all_param_cols)

    conflict_cells = len(conflicts_detail)
    conflict_instances = len(conflict_rows)

    phase2_report = run_phase2_checks(
        wb,
        parse_results=parse_results,
        control_ids={r["cell_id"] for r in control_rows},
        effect_rows=effect_rows,
        label_merge_count=len(merge_log),
        conflict_cell_count=conflict_cells,
        cluster_fill_count=cluster_fills,
        excluded_in_may_count=len(excluded_may),
        exclude_flag_dropped=len(exclude_flag),
    )

    phase3_report = run_phase3_checks(
        parse_results=parse_results,
        control_rows=control_rows,
        effect_rows=effect_rows,
        param_cols=param_cols,
        dropped_ids=dropped_ids,
        all_cells=all_cells,
        clusters=clusters,
        tags=tags,
    )

    skip_format = meta_set | {"experiment", "conflict_source_sheet"}
    _format_all_row_params(control_rows, skip_keys=skip_format)
    _format_all_row_params(effect_rows, skip_keys=skip_format)
    _format_all_row_params(conflict_rows, skip_keys=set(CONFLICT_META_COLUMNS))

    control_cols = META_COLUMNS + param_cols
    effect_cols = META_COLUMNS + ["experiment"] + param_cols
    conflict_param_cols = _ordered_param_columns(all_conflict_param_cols)
    conflict_cols = CONFLICT_META_COLUMNS + conflict_param_cols

    output_dir.mkdir(parents=True, exist_ok=True)
    _write_workbook(
        output_dir / "Intrinsic_master.xlsx",
        [
            ("control_excitability", control_rows, control_cols),
            ("pharmacology_effect", effect_rows, effect_cols),
            ("duplicate_conflicts", conflict_rows, conflict_cols),
        ],
    )
    _write_csv(output_dir / "control_excitability.csv", control_rows, control_cols)
    _write_csv(output_dir / "pharmacology_effect.csv", effect_rows, effect_cols)
    _write_csv(output_dir / "duplicate_conflicts.csv", conflict_rows, conflict_cols)

    phase4_report = write_qc_workbook(
        output_dir,
        control_rows=control_rows,
        effect_rows=effect_rows,
        meta_columns=META_COLUMNS,
        param_columns=param_cols,
    )

    report: dict[str, Any] = {
        "source": str(source_path),
        "control_neurons": len(control_rows),
        "effect_rows": len(effect_rows),
        "conflict_rows": conflict_instances,
        "conflict_cells": conflict_cells,
        "excluded_in_may_dropped": len(excluded_may),
        "exclude_flag_dropped": len(exclude_flag),
        "dedup_priority_note": "All Analysed data is canonical when overlapping with area sheets",
        "label_merges": merge_log,
        "cluster_fills_assumed_type": cluster_fills,
        "cluster_fill_column": "assumed_type",
        "param_columns": len(param_cols),
        "duplicate_header_warnings": header_warnings,
        "duplicate_conflicts_detail": conflicts_detail,
        "region_area_conflicts": region_conflicts,
        "phase2_verification": phase2_report,
        "phase3_verification": phase3_report,
        "phase4_qc": phase4_report,
    }
    manifest = write_run_manifest(output_dir, source_path=source_path, report=report)

    wb.close()
    return manifest


def _write_csv(path: Path, rows: list[dict], columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow(row)


def _write_workbook(
    path: Path,
    sheets: list[tuple[str, list[dict], list[str]]],
) -> None:
    out = Workbook()
    out.remove(out.active)
    for title, rows, columns in sheets:
        ws = out.create_sheet(title)
        ws.append(columns)
        ws.freeze_panes = "A2"
        for row in rows:
            ws.append([row.get(c) for c in columns])
    out.save(path)
