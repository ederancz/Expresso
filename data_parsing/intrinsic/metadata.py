"""Metadata joins from All cells and metadata-only sheets."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook

from .area import normalize_region, resolve_area
from .config import ALL_CELLS_SHEET, CLUSTER_SHEET, METADATA_SHEETS
from .ids import normalize_id


@dataclass
class CellMetadata:
    source_sheet: str = ""
    region: str = ""
    region_sheet: str = ""
    region_conflict: bool = False
    layer: str = ""
    projection_target: str = ""
    classic_burster: Any = None
    area_ccf: str = ""
    exclude_flag: Any = None
    cre_label: Any = None
    axon: Any = None
    notes: str = ""
    time_from_5ht: Any = None
    assumed_type: str = ""
    physiological_cluster: str = ""
    dup_conflict: bool = False


def _normalize_pt(val: str) -> str:
    return "ET" if val.strip().upper() == "PT" else val


def _output_assumed_type(val: str) -> str:
    if not val:
        return ""
    val = _normalize_pt(val)
    if val == "Tlx":
        return "IT"
    return val


def _merge_notes(note: Any, comment: Any) -> str:
    parts: list[str] = []
    for val in (note, comment):
        if val is None:
            continue
        text = str(val).strip()
        if text:
            parts.append(text)
    return " | ".join(parts)


def is_excluded_in_may(val: Any) -> bool:
    return val in (1, 1.0, "1", True)


def is_exclude_flag(val: Any) -> bool:
    """True when All cells exclude_flag marks the cell for drop at parse time."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return int(val) == 1
    return str(val).strip() in {"1", "1.0", "True", "true"}


def load_excluded_may_ids(wb: Workbook) -> set[str]:
    ws = wb[ALL_CELLS_SHEET]
    out: set[str] = set()
    for r in range(1, ws.max_row + 1):
        cid_raw = ws.cell(r, 1).value
        if cid_raw is None:
            continue
        cid = normalize_id(str(cid_raw).strip())
        if not cid.startswith("nm"):
            continue
        if is_excluded_in_may(ws.cell(r, 14).value):
            out.add(cid)
    return out


def load_exclude_flag_ids(wb: Workbook) -> set[str]:
    ws = wb[ALL_CELLS_SHEET]
    out: set[str] = set()
    for r in range(1, ws.max_row + 1):
        cid_raw = ws.cell(r, 1).value
        if cid_raw is None:
            continue
        cid = normalize_id(str(cid_raw).strip())
        if not cid.startswith("nm"):
            continue
        if is_exclude_flag(ws.cell(r, 2).value):
            out.add(cid)
    return out


def load_all_cells(wb: Workbook) -> dict[str, dict[str, Any]]:
    ws = wb[ALL_CELLS_SHEET]
    out: dict[str, dict[str, Any]] = {}
    for r in range(1, ws.max_row + 1):
        cid_raw = ws.cell(r, 1).value
        if cid_raw is None:
            continue
        cid = normalize_id(str(cid_raw).strip())
        if not cid.startswith("nm"):
            continue
        out[cid] = {
            "exclude_flag": ws.cell(r, 2).value,
            "time_from_5ht": ws.cell(r, 4).value,
            "cre_label": ws.cell(r, 7).value,
            "all_cells_area": ws.cell(r, 8).value,
            "axon": ws.cell(r, 9).value,
            "note": ws.cell(r, 10).value,
            "excluded_in_may": ws.cell(r, 14).value,
            "comment": ws.cell(r, 15).value,
        }
    return out


def load_metadata_tags(wb: Workbook) -> dict[str, dict[str, str]]:
    tags: dict[str, dict[str, str]] = {}
    for sheet_name, tag_values in METADATA_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is None:
                continue
            s = str(v).strip()
            if not (s.lower().startswith("nm") or __import__("re").match(r"^\d{4}_", s)):
                continue
            cid = normalize_id(s)
            tags.setdefault(cid, {}).update(tag_values)
    return tags


def load_cluster_metadata(wb: Workbook) -> dict[str, str]:
    if CLUSTER_SHEET not in wb.sheetnames:
        return {}
    ws = wb[CLUSTER_SHEET]
    out: dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        cluster = ws.cell(r, 9).value
        cid_raw = ws.cell(r, 10).value
        if cid_raw is None or cluster is None:
            continue
        cid = normalize_id(str(cid_raw).strip())
        out[cid] = str(cluster).strip()
    return out


def fill_assumed_type_from_cluster(meta: CellMetadata) -> bool:
    if meta.assumed_type or not meta.physiological_cluster:
        return False
    cl = meta.physiological_cluster
    if cl == "IT":
        meta.assumed_type = "IT"
    elif cl in ("ET1", "ET2"):
        meta.assumed_type = "ET"
    return True


def merge_cell_metadata(
    cell_id: str,
    *,
    source_sheet: str,
    region: str,
    layer: str,
    sheet_meta: dict[str, Any],
    all_cells: dict[str, dict[str, Any]],
    tags: dict[str, dict[str, str]],
    clusters: dict[str, str],
) -> CellMetadata:
    ac = all_cells.get(cell_id, {})
    morph_raw = sheet_meta.get("area_morph_raw", {}).get(cell_id)
    broad_area, area_ccf, _ = resolve_area(
        sheet_region=region,
        morph_raw=morph_raw,
        all_cells_area=ac.get("all_cells_area"),
        all_cells_note=ac.get("note"),
    )

    region_sheet = normalize_region(region)
    final_region = broad_area or region_sheet
    region_conflict = bool(broad_area and region_sheet and broad_area != region_sheet)

    meta = CellMetadata(
        source_sheet=source_sheet,
        region=final_region,
        region_sheet=region_sheet,
        region_conflict=region_conflict,
        layer=layer,
        classic_burster=sheet_meta.get("classic_burster", {}).get(cell_id),
        area_ccf=area_ccf,
        exclude_flag=ac.get("exclude_flag"),
        cre_label=ac.get("cre_label"),
        axon=ac.get("axon"),
        notes=_merge_notes(ac.get("note"), ac.get("comment")),
        time_from_5ht=ac.get("time_from_5ht"),
    )

    for k, v in tags.get(cell_id, {}).items():
        if k == "assumed_type":
            meta.assumed_type = v
        else:
            setattr(meta, k, v)

    if cell_id in clusters:
        meta.physiological_cluster = clusters[cell_id]

    fill_assumed_type_from_cluster(meta)
    meta.assumed_type = _output_assumed_type(meta.assumed_type)
    return meta


def metadata_to_row(meta: CellMetadata) -> dict[str, Any]:
    return {
        "cell_id": "",
        "region": meta.region,
        "areaCCF": meta.area_ccf,
        "layer": meta.layer,
        "projection_target": meta.projection_target,
        "assumed_type": meta.assumed_type,
        "physiological_cluster": meta.physiological_cluster,
        "source_sheet": meta.source_sheet,
        "classic_burster": meta.classic_burster,
        "exclude_flag": meta.exclude_flag,
        "cre_label": meta.cre_label,
        "axon": meta.axon,
        "notes": meta.notes,
        "time_from_5HT": meta.time_from_5ht,
        "dup_conflict": meta.dup_conflict,
    }
